import argparse
import re
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, numbers
import logging

# import numpy as np


parser = argparse.ArgumentParser(description="SRT to excel for translators")

parser.add_argument(
    "-ilang", "--inputlang", type=str, metavar="", required=False, help="input language"
)

parser.add_argument(
    "-fix_c",
    "--fix_comma",
    type=str,
    metavar="",
    required=False,
    help="whether to fix commas or not",
)

logging.basicConfig(
    filename=r"C:\Users\flitto\Desktop\vswd\logfiles\ali_excel_txt.log",
    level=logging.INFO,
)


# def frame_converter(arrows_del_timecode):
#     start, mili = str(arrows_del_timecode).split(",")
#     ff = "0." + mili[0:]
#     ff = float(ff)
#     ff = ff * 24.0
#     ff = round(ff, 0)
#     if ff == 24.0:
#         ff = 0.0
#     else:
#         pass
#     ff = str(ff)
#     dotloc = ff.rfind(".")
#     ff = ff[:dotloc]
#     ff = np.char.zfill(ff, 2)
#     final_time = str(start) + ":" + str(ff)
#     return final_time


def remove_empty_break(listlike):  # \n 밖에 없는 라인들 처리.
    while True:
        try:
            listlike.remove("\n")  # remove(x) method 는 리스트 내의 가장 앞에 있는 x 값을 지워준다.
        except ValueError:
            break
    return listlike


def pop_empty_break(listlike):
    for ind, n_break in enumerate(listlike):
        if n_break == "\n":
            listlike.pop(ind)  # .pop
        else:
            pass
    return listlike


def check_line_number(n1, n2, fnames):  # 숫자가 1씩 추가되는가?
    n1 = int(n1)
    n2 = int(n2)
    if n1 + 1 != n2:
        logging.warning(
            f"Operation: Excelify      File: {fnames: <40}  LINE: {n1: <5} and {n2: <5} ERROR TYPE: Line Number"
        )
    else:
        pass


def remove_breaks(str_or_list):  # 각 줄마다 있는 \n 제거
    if isinstance(str_or_list, str):
        script = re.sub("[\n]", "", str_or_list)  # 한 숫자에 한줄의 transcription만 있으면 그냥 제거
        return script
    else:
        return "".join([re.sub("[\n]", " ", x) for x in str_or_list])


def line_number_index(empty_linebreak_removed):

    line_location_list = []
    pattern = re.compile("[\d]+[\s]")

    for n, lin in enumerate(empty_linebreak_removed):
        # verifying the format of srt, int and then timecode -> int = line number
        if re.fullmatch(pattern, lin) and re.search(
            "-->", empty_linebreak_removed[n + 1]
        ):
            line_location_list.append(n)
        else:
            pass
    return line_location_list


def remove_minus(s):
    if s.startswith("-"):
        if s.startswith("--"):
            s = s[2:]
        else:
            s = s[1:]
    else:
        pass
    return s


def srt_list(no_breaks, no_breaks_index, f):
    final_list = []  # 자막 행 리스트들의 리스트
    for i in range(len(no_breaks_index)):  # srt 파일내 라인 번호의 위치 기준으로 작업
        parsed_srt = []  # 자막 행 리스트 순서는 line number, 텍스트, 타임코드1, 타임코드2

        ind_line_loc = no_breaks_index[i]  # same as for i in no_breaks_index
        parsed_srt.append(remove_breaks(str(no_breaks[ind_line_loc])))  # line number

        ind_tc = (
            ind_line_loc + 1
        )  # 타임코드의 인덱스는 linenumber 인덱스의 다음. (verified by line_number_index())
        arrows_timecode = remove_breaks(no_breaks[ind_tc])
        beginning, end = str(arrows_timecode).split("-->")  # 타임코드 처리
        beginning = str(beginning).strip()
        end = str(end).strip()

        # srt_start = frame_converter(beginning.strip())
        # srt_end = frame_converter(end.strip())

        ind_script_start = ind_tc + 1  # 스크립트는 타임코드 다음 인덱스
        if i + 1 >= len(
            no_breaks_index
        ):  # 마지막 loop 돌때 마지막 transcription 추출용 if statement
            pre_script = str(
                no_breaks[ind_script_start]
            )  # ind _script_start이지만 마지막 돌때는 이게 마지막 칸이다.
            worked_script = remove_breaks(pre_script)
            # worked_script = remove_minus(worked_script)
            parsed_srt.append(worked_script)

        else:
            ind_script_end = no_breaks_index[i + 1]
            check_line_number(
                no_breaks[ind_line_loc], no_breaks[ind_script_end], f
            )  # line number 가 맞는지 확인, 틀릴시 로깅
            pre_script = no_breaks[
                ind_script_start:ind_script_end
            ]  # 두줄 + 있는 transcrived line 들을 위해
            worked_script = remove_breaks(pre_script).strip()
            # worked_script = remove_minus(worked_script)
            parsed_srt.append(worked_script)

        parsed_srt.append(beginning)
        parsed_srt.append(end)

        final_list.append(parsed_srt)
    return final_list


def srt_to_df(srt_list, input_lang, output_lang):
    cols = [
        "Lang",
        "Line 1",
        "Start time (hh:mm:ss,ms)",
        "End time (hh:mm:ss,ms)",
        "Lang1",
        "Translation1",
    ]

    excel_df = pd.DataFrame(srt_list)
    excel_df[0] = input_lang
    print(output_lang)
    if len(output_lang) > 4:
        cols = [
            "Lang",
            "Line 1",
            "Start time (hh:mm:ss,ms)",
            "End time (hh:mm:ss,ms)",
            "Lang1",
            "Translation1",
            "Lang2",
            "Translation2",
            "Lang3",
            "Translation3",
        ]

        output_lang1 = output_lang[0:2]
        output_lang2 = output_lang[2:4]
        output_lang3 = output_lang[4:]
        excel_df.insert(len(excel_df.columns), "lang1", output_lang1)
        excel_df.insert(len(excel_df.columns), "Translation1", "")

        excel_df.insert(len(excel_df.columns), "lang2", output_lang2)
        excel_df.insert(len(excel_df.columns), "Translation2", "")

        excel_df.insert(len(excel_df.columns), "lang3", output_lang3)
        excel_df.insert(len(excel_df.columns), "Translation3", "")
        excel_df.columns = cols
    elif len(output_lang) <= 4 and len(output_lang) > 2:
        cols = [
            "Lang",
            "Line 1",
            "Start time (hh:mm:ss,ms)",
            "End time (hh:mm:ss,ms)",
            "Lang1",
            "Translation1",
            "Lang2",
            "Translation2",
        ]

        output_lang1 = output_lang[0:2]
        output_lang2 = output_lang[2:]
        excel_df.insert(len(excel_df.columns), "lang1", output_lang1)
        excel_df.insert(len(excel_df.columns), "Translation1", "")

        excel_df.insert(len(excel_df.columns), "lang2", output_lang2)
        excel_df.insert(len(excel_df.columns), "Translation2", "")

        excel_df.columns = cols
    else:
        excel_df.insert(len(excel_df.columns), "lang", output_lang)
        excel_df.insert(len(excel_df.columns), "Translation", "")
        excel_df.columns = cols
    return excel_df


def verify_srtdf(df, col, fname):  # 받는 값은 df, 확인할 열 index, 그리고 로깅을 위한 파일명
    verify_bool = df[df.columns[col]] == ""
    if len(df[verify_bool]) > 0:
        wrong_lines = df[verify_bool].index.tolist()
        for empty in wrong_lines:
            empty = empty + 1
            logging.warning(
                f"Operation: Excelify      File: {fname: <40}  LINE: {empty: <10} ERROR TYPE: Empty Cell"
            )
    else:
        pass


def fix_comma_ends(df):
    final_lines = []
    # chinese_commas = ['，','、']
    check_lines = df["Line 1"].tolist()
    tc_beg = df[df.columns[2]].tolist()
    tc_end = df[df.columns[3]].tolist()
    span_lines = ""
    for n, line in enumerate(check_lines):
        line = ("").join(line.split())
        if span_lines == "":
            start_tc = tc_beg[n]
        if line.endswith("，") or line.endswith("、"):
            span_lines += line
        else:
            end_tc = tc_end[n]
            span_lines += line
            final_lines.append([span_lines, start_tc, end_tc])
            span_lines = ""
    return final_lines


def final_fix_df(df):
    final_df = pd.DataFrame(data=fix_comma_ends(df), columns=df.columns[1:4])
    final_df.insert(0, "Lang", df["Lang"].iloc[0])
    for x in range(4, len(df.columns)):
        final_df[df.columns[x]] = df[df.columns[x]].iloc[0]
    return final_df


def tc_reformat(tc, fname):
    tc = ("").join(tc.split())
    end_col = tc.rfind(":")

    if len(tc[end_col + 1 :]) > 3:
        pass
    else:
        tc = tc[:end_col] + "," + tc[end_col + 1 :]
    tc = time_code_check(tc, fname)

    return tc


def time_code_check(tc, file_name):
    tc = ("").join(str(tc).split())
    try:
        if tc[-4] != "," or len(tc) != 12:
            logging.warning(
                f"Operation: Textify       File: {file_name}: {tc: <30} ERROR TYPE: tc error"
            )
            # raise ValueError("Timecode in wrong format.")
    except:
        logging.warning(
            f"Operation: Textify       File: {file_name}: {tc: <30} ERROR TYPE: tc error"
        )
        # raise ValueError("Timecode in wrong format.")
    return tc


def replace_for_comparison(tc):
    tc = tc.replace(":", "")
    tc = tc.replace(",", ".")
    tc = float(tc)
    return tc


def tc_qc(timecode_beg, timecode_end, file_name):
    timecode_beg = timecode_beg.tolist()
    timecode_end = timecode_end.tolist()
    for n, b in enumerate(timecode_beg):
        e = timecode_end[n]
        if b.find(",") != -1 and e.find(",") != -1:
            if n == 0:
                fixed_prev_e = float(0)
            else:
                prev_e = timecode_end[n - 1]
                fixed_prev_e = replace_for_comparison(prev_e)
            fixed_b = replace_for_comparison(b)
            fixed_e = replace_for_comparison(e)
            try:
                if fixed_b >= fixed_e:  # 현재 타임코드의 스타트가 엔드보다 클때 (늦을때)
                    logging.warning(
                        f"Operation: Excelify       File: {file_name: <40} ERROR TYPE: timecode1 {n+1}  {b, e}"
                    )
                if fixed_prev_e > fixed_b:  # 현재 타임코드가 이전 타임코드보다 적을때 (이를때)
                    logging.warning(
                        f"Operation: Excelify       File: {file_name: <40} ERROR TYPE: timecode2 {n+1}  {fixed_prev_e, fixed_b}"
                    )
            except:  # 전반적인 타임코드 에러
                logging.warning(
                    f"Operation: Excelify       File: {file_name: <40} ERROR TYPE: timecode3 {n+1}  {b, e}"
                )
        else:  # 전반적인 타임코드 에러
            logging.warning(
                f"Operation: Excelify       File: {file_name: <40} ERROR TYPE: timecode3 {n+1}  {b, e}"
            )


def srt_to_excel(df, splitlen, dst, fnames, fix_comma):
    count = round(len(df) / splitlen)  # to_excel 을 1500줄 이상 나눠준다. splitlen은 1000,
    # 1500/1000 을 반올림하면, 그때부터 count는 2가 되기 때문.
    if fix_comma == "fix":
        df = final_fix_df(df)
    if len(df) > splitlen and count > 1:
        split_keys = ["？", "?", ".", "!", "。", "!", "#"]
        splits = round(len(df) / count)
        i = 1
        start = 0
        while count > 1:
            out_name = f"{fnames}_{i}"
            for n, x in enumerate(df[df.columns[1]].iloc[splits * i :]):
                if x[-1] in split_keys and (((splits * i) + n) - start) > splits:
                    split_pos = n + 1 + (splits * i)
                    split_df = df.iloc[start:split_pos]
                    verify_srtdf(split_df, 1, str(out_name + ".xlsx"))
                    split_df.to_excel(
                        f"{dst}\\{out_name}.xlsx",
                        sheet_name=f"Sheet 1",
                        index=False,
                        encoding="utf-8-sig",
                    )
                    break
            start = split_pos
            i += 1
            count -= 1

        out_name = f"{fnames}_{i}"
        split_df = df.iloc[start:]
        verify_srtdf(split_df, 1, str(out_name + ".xlsx"))

        df[df.columns[2]] = df[df.columns[2]].apply(tc_reformat, args=(out_name,))
        df[df.columns[3]] = df[df.columns[3]].apply(tc_reformat, args=(out_name,))
        tc_qc(df[df.columns[2]], df[df.columns[3]], out_name)

        split_df.to_excel(
            f"{dst}\\{out_name}.xlsx",
            sheet_name=f"Sheet 1",
            index=False,
            encoding="utf-8-sig",
        )
    else:
        verify_srtdf(df, 1, str(fnames + ".xlsx"))

        df[df.columns[2]] = df[df.columns[2]].apply(tc_reformat, args=(fnames,))
        df[df.columns[3]] = df[df.columns[3]].apply(tc_reformat, args=(fnames,))
        tc_qc(df[df.columns[2]], df[df.columns[3]], fnames)

        df.to_excel(
            f"{dst}\\{fnames}.xlsx",
            sheet_name=f"Sheet 1",
            index=False,
            encoding="utf-8-sig",
        )


def format_excel(excel_file, final_path):
    filepath = os.path.join(final_path, excel_file)

    wb = openpyxl.load_workbook(f"{filepath}")
    ws = wb.worksheets[0]

    bold_font = Font(bold=False)
    side = openpyxl.styles.Side(border_style=None)
    no_border = openpyxl.styles.borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )

    for cell in ws["1:1"]:
        cell.font = bold_font
        cell.border = no_border
    for acell in ws["B"]:
        acell.number_format = numbers.FORMAT_TEXT

    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["F"].width = 55
    if ws["H2"] == "EN" or "FR":
        ws.column_dimensions["H"].width = 55
    if ws["J2"] == "PT" or ws["J2"] == "FR":
        ws.column_dimensions["J"].width = 55
    wb.save(f"{filepath}")


if __name__ == "__main__":
    args = parser.parse_args()
    workinput = args.inputlang + "\\work"  # 폴더 구조 바꿈
    done = args.inputlang + "\\done"
    fix_comma = args.fix_comma
    cwd = r"C:\Users\flitto\Desktop\vswd\alifiles"
    fwd = r"C:\Users\flitto\Desktop\vswd\output_files\ali_out\current"
    wd = os.path.join(cwd, workinput)
    rename_dir = os.path.join(cwd, done)
    langcode_dict = {
        "ZH": "ENFRPT",
        "EN": "FRPT",
        "ES": "EN",
        "FR": "ENPT",
    }  # input 언어 기준으로
    # output 언어를 잡아준다. 그에 따른 작업할 폴더 경로도 획득.

    output_lang = langcode_dict.get(args.inputlang)

    for pre_srt in os.listdir(wd):
        srt_file = open(wd + "\\" + pre_srt, "r", encoding="utf-8-sig")
        lines = srt_file.readlines()
        srt_file.close()

        srt_name = pre_srt[: pre_srt.rfind(".")]
        newname = str(pre_srt[: pre_srt.rfind("_")]) + f"_{args.inputlang}{output_lang}"

        line_break_fixed = remove_empty_break(lines)
        line_indexes = line_number_index(line_break_fixed)
        srt = srt_list(line_break_fixed, line_indexes, pre_srt)
        if len(output_lang) > 2:
            output_lang1 = output_lang[0:2]
            output_lang2 = output_lang[2:]
            srt_df1 = srt_to_df(srt, args.inputlang, output_lang)
            newname1 = (
                str(pre_srt[: pre_srt.rfind("_")])
                + f"_{args.inputlang}{output_lang1}{output_lang2}"
            )
            # srt_df2 = srt_to_df(srt, args.inputlang, output_lang2)
            # newname2 = (
            #     str(pre_srt[: pre_srt.rfind("_")]) + f"_{args.inputlang}{output_lang2}"
            # )
            srt_to_excel(srt_df1, 100000, fwd, newname1, fix_comma)
            # srt_to_excel(srt_df2, 100000, fwd, newname2)
            format_excel(newname1 + ".xlsx", fwd)
            # format_excel(newname2 + ".xlsx", fwd)
        else:
            srt_df = srt_to_df(srt, args.inputlang, output_lang)

            srt_to_excel(srt_df, 100000, fwd, newname, fix_comma)
            format_excel(newname + ".xlsx", fwd)
        try:
            os.rename(os.path.join(wd, pre_srt), os.path.join(rename_dir, pre_srt))
        except:
            print("alr in folder")
